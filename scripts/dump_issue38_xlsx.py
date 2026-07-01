#!/usr/bin/env python3
"""Dump every row of the issue #38 Sloka Master List Excel.

Prints all columns so the full change list is visible.
Usage: python3 scripts/dump_issue38_xlsx.py [path/to/xlsx]
"""
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"], check=True)
    from openpyxl import load_workbook


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else "scratch/sloka_master.xlsx"
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        print(f"\n===== SHEET: {ws.title}  (dims={ws.dimensions}) =====")
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print("(empty)")
            continue
        header = rows[0]
        print("HEADER:", header)
        for i, row in enumerate(rows[1:], start=2):
            # Skip fully-empty rows
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            cells = [("" if c is None else str(c)) for c in row]
            print(f"--- row {i} ---")
            for h, c in zip(header, cells):
                print(f"    [{h!r}] = {c!r}")


if __name__ == "__main__":
    main()
