#!/usr/bin/env python3
"""Extract only the actionable rows (with a change instruction) from the
issue #38 Sloka Master List. Outputs JSON to scratch/issue38_changes.json.
"""
import json
import sys

from openpyxl import load_workbook


def norm(c):
    return "" if c is None else str(c).strip()


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else "scratch/sloka_master.xlsx"
    wb = load_workbook(path, data_only=True)
    ws = wb["Detailed List"] if "Detailed List" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    # Use the first 4 columns per spec
    out = []
    for i, row in enumerate(rows[1:], start=2):
        chapter = norm(row[0])
        sloka = norm(row[1])
        padas = norm(row[2])
        desc = norm(row[3])
        if not padas and not desc:
            continue  # no change per spec
        if not chapter and not sloka:
            continue
        out.append({
            "row": i,
            "chapter": chapter,
            "sloka": sloka,
            "padas": padas,
            "desc": desc,
        })
    print(f"Total actionable rows: {len(out)}")
    # distinct chapters
    chaps = {}
    for o in out:
        chaps.setdefault(o["chapter"], 0)
        chaps[o["chapter"]] += 1
    print("By chapter:", json.dumps(chaps, ensure_ascii=False))
    with open("scratch/issue38_changes.json", "w") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print("Wrote scratch/issue38_changes.json")


if __name__ == "__main__":
    main()
